#!/usr/bin/env python3
"""
Sistema Predictivo de Unfullfilment - UNICEF Argentina 2026
Entrena un modelo de regresion logistica con DIC-2025/ENE-2026 (labels ya definidos)
y scoreea los donantes activos MAR-MAY-2026.
Genera un dashboard HTML standalone.
"""

import json, math, os, unicodedata, warnings
from collections import defaultdict
from datetime import datetime

import openpyxl
import pandas as pd
import numpy as np
from sklearn.linear_model import LogisticRegression
from sklearn.preprocessing import StandardScaler
from sklearn.pipeline import Pipeline
from sklearn.metrics import roc_auc_score
from sklearn.model_selection import cross_val_score

warnings.filterwarnings("ignore")

def strip_accents(s: str) -> str:
    return unicodedata.normalize("NFD", s).encode("ascii", "ignore").decode()

BASE     = os.path.dirname(os.path.abspath(__file__))
FILE_DON = os.path.join(BASE, "CENTRO DE COMANDO UNICEF ARGENTINA 2026 - DONACIONES (1).csv")
FILE_UF  = os.path.join(BASE, "CENTRO DE COMANDO UNICEF ARGENTINA 2026 - UNFULLFILMENT (1).csv")
FILE_NOM = os.path.join(os.path.dirname(BASE), "Copia de NOMINA PROA 2024 - para acciones.xlsx")
OUTPUT   = os.path.join(BASE, "predictivo_uf_dashboard.html")

# ── 0. NOMINA: estado de cada facer por username ──────────────────────────────
print("Leyendo nomina Argentina...")

def _load_nomina_status(path: str) -> dict:
    """Returns {username_lower: {'estado': 'ACTIVO'|'BAJA', 'puesto': str, 'meses_prov': [...]}}"""
    status: dict = {}
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

        # PROV. USUARIO EG: facers que cobraron en DIC-2025/ENE-2026/FEB-2026
        prov_by_mes: dict = defaultdict(set)
        if "PROV. USUARIO EG" in wb.sheetnames:
            ws_prov = wb["PROV. USUARIO EG"]
            for row in ws_prov.iter_rows(values_only=True):
                if not row or len(row) < 7:
                    continue
                mes = str(row[1] or "").strip()
                u   = str(row[6] or "").strip().lower()
                if mes and u and mes not in ("MES-AÑO", "MES-ANO", ""):
                    prov_by_mes[mes].add(u)

        meses_recientes = ["DIC-2025", "ENE-2026", "FEB-2026"]
        en_prov = set()
        for m in meses_recientes:
            en_prov |= prov_by_mes.get(m, set())

        for u in en_prov:
            meses = [m for m in meses_recientes if u in prov_by_mes.get(m, set())]
            status[u] = {"estado": "ACTIVO", "puesto": "Facer", "meses_prov": meses}

        # NOMINA ARGENTINA: complementar puesto y marcar bajas
        if "NOMINA ARGENTINA" in wb.sheetnames:
            ws_nom = wb["NOMINA ARGENTINA"]
            rows_nom = list(ws_nom.iter_rows(values_only=True))
            for row in rows_nom[2:]:
                if not any(v is not None for v in row):
                    continue
                ev_id  = str(row[44] or "").strip().lower()
                estado = str(row[7]  or "").strip()
                puesto = str(row[8]  or "").strip()
                if not ev_id or ev_id in ("-", "none") or ev_id.replace(".", "").isdigit():
                    continue
                if ev_id in status:
                    status[ev_id]["puesto"] = puesto
                elif estado == "BAJA":
                    status[ev_id] = {"estado": "BAJA", "puesto": puesto, "meses_prov": []}
        wb.close()
    except Exception as e:
        print(f"  Advertencia nomina: {e}")
    return status

facer_status = _load_nomina_status(FILE_NOM)
n_activos = sum(1 for v in facer_status.values() if v["estado"] == "ACTIVO")
n_bajas   = sum(1 for v in facer_status.values() if v["estado"] == "BAJA")
print(f"  Activos en nomina: {n_activos} | Bajas en nomina: {n_bajas}")

# ── 1. CARGA ──────────────────────────────────────────────────────────────────
print("Cargando archivos...")
df = pd.read_csv(FILE_DON, encoding="utf-8-sig", dtype=str).fillna("")
uf = pd.read_csv(FILE_UF,  encoding="utf-8-sig", dtype=str).fillna("")
# Normalize: strip whitespace + remove accents from column names
df.columns = [strip_accents(c.strip()) for c in df.columns]
uf.columns = [strip_accents(c.strip()) for c in uf.columns]
# Normalize string values in df (strip whitespace only; keep values legible)
for col in df.columns:
    df[col] = df[col].str.strip()

# Solo filas con mes asignado  (column now: MES-ANO after normalization)
df = df[df["MES-ANO"] != ""].copy()
print(f"  Donaciones con mes: {len(df)}")

# ── 2. TASA UF HISTORICA POR FACER (username) ─────────────────────────────────
facer_stats = defaultdict(lambda: {"t": 0, "uf": 0})
for _, r in df.iterrows():
    u = str(r["Usuario"])
    facer_stats[u]["t"] += 1
    if str(r["UNFULLFILMENT"]) == "SI":
        facer_stats[u]["uf"] += 1

GLOBAL_UF_RATE = (df["UNFULLFILMENT"] == "SI").sum() / len(df)
print(f"  Tasa UF global: {GLOBAL_UF_RATE:.3f}")

LAPLACE_N = 10
def facer_uf_rate(usuario: str) -> float:
    s = facer_stats[str(usuario).strip()]
    return (s["uf"] + GLOBAL_UF_RATE * LAPLACE_N) / (s["t"] + LAPLACE_N)

# ── 3. FEATURE ENGINEERING ────────────────────────────────────────────────────
NOMBRE_FEAT = {
    "log_monto":   "Monto de donacion (log)",
    "email":       "Tiene email",
    "celular":     "Tiene celular",
    "domicilio":   "Tiene domicilio",
    "tel_fijo":    "Tiene telefono fijo",
    "edad_25_34":  "Edad 25-34",
    "edad_menos25":"Edad < 25",
    "edad_35_44":  "Edad 35-44",
    "edad_45_54":  "Edad 45-54",
    "edad_55_64":  "Edad 55-64",
    "edad_65_74":  "Edad 65-74",
    "tc_debito":   "Tarjeta debito",
    "tc_credito":  "Tarjeta credito",
    "marca_mc":    "Marca Mastercard",
    "marca_visa":  "Marca VISA",
    "marca_naranja":"Marca Naranja",
    "marca_cabal": "Marca Cabal",
    "marca_amex":  "Marca Amex",
    "aac_acepta":  "Acepta AAC",
    "gen_masc":    "Genero masculino",
    "quincena_1ra":"1ra quincena",
    "cuarenta":    "40+ anios",
}
FEATURE_COLS = list(NOMBRE_FEAT.keys())

df["monto_num"] = pd.to_numeric(
    df["Monto de donacion"].str.replace(r"[^\d]", "", regex=True), errors="coerce"
).fillna(0)
# normalize Tipo de Tarjeta values (strip accents for comparison)
df["tc_norm"] = df["Tipo de Tarjeta"].apply(strip_accents)

def encode_row(r) -> list:
    edad  = str(r.get("Rango etario", ""))
    # use tc_norm (accent-stripped) for comparison
    tc    = str(r.get("tc_norm", strip_accents(str(r.get("Tipo de Tarjeta", "")))))
    marca = str(r.get("Marca de Tarjeta", ""))
    aac   = str(r.get("AAC", ""))
    mn    = float(r.get("monto_num", 0) or 0)
    tel   = str(r.get("Telefono Fijo - Laboral - Otro - Alternativo", "")).upper()
    return [
        math.log1p(mn),
        1 if str(r.get("Email",    "")).upper() == "SI" else 0,
        1 if str(r.get("Celular",  "")).upper() == "SI" else 0,
        1 if str(r.get("Domicilio","")).upper() == "SI" else 0,
        1 if tel == "SI" else 0,
        1 if edad == "25 a 34"  else 0,
        1 if edad == "< 25"     else 0,
        1 if edad == "35 a 44"  else 0,
        1 if edad == "45 a 54"  else 0,
        1 if edad == "55 a 64"  else 0,
        1 if edad == "65 a 74"  else 0,
        1 if tc == "Debito"     else 0,
        1 if tc == "Credito"    else 0,
        1 if marca == "MASTERCARD"      else 0,
        1 if marca == "VISA"            else 0,
        1 if marca == "TARJETA NARANJA" else 0,
        1 if marca == "CABAL"           else 0,
        1 if marca == "AMEX"            else 0,
        1 if "Acepta" in aac    else 0,
        1 if str(r.get("Genero del donante", "")) == "Masculino" else 0,
        1 if "1ra" in str(r.get("QUINCENA", "")) else 0,
        1 if str(r.get("40  ANOS", "")).upper() == "SI" else 0,
    ]

# ── 4. ENTRENAMIENTO: DIC-2025 + ENE-2026 ────────────────────────────────────
print("Construyendo dataset de entrenamiento (DIC-2025 + ENE-2026)...")
df_train = df[df["MES-ANO"].isin({"DIC-2025", "ENE-2026"})].copy()
pos = (df_train["UNFULLFILMENT"] == "SI").sum()
print(f"  Filas: {len(df_train)} | UF=SI: {pos} ({pos/len(df_train)*100:.1f}%)")

X_train = np.array([encode_row(r) for _, r in df_train.iterrows()])
y_train = np.where(df_train["UNFULLFILMENT"] == "SI", 1, 0)

pipe = Pipeline([
    ("scaler", StandardScaler()),
    # sin class_weight para obtener probabilidades calibradas (mediana ~tasa base real)
    ("clf",    LogisticRegression(max_iter=2000, C=0.5, solver="lbfgs")),
])
pipe.fit(X_train, y_train)

cv_auc  = cross_val_score(pipe, X_train, y_train, cv=5, scoring="roc_auc")
train_auc = roc_auc_score(y_train, pipe.predict_proba(X_train)[:, 1])
print(f"  Train AUC: {train_auc:.3f} | CV AUC: {cv_auc.mean():.3f} +- {cv_auc.std():.3f}")

coefs = pipe.named_steps["clf"].coef_[0]
importances = sorted(zip(FEATURE_COLS, coefs), key=lambda x: abs(x[1]), reverse=True)
print("  Top features:", [(n, round(c, 3)) for n, c in importances[:6]])

# ── 5. SCORING: donantes UNFULLFILMENT=NO ─────────────────────────────────────
print("Scorando donantes activos...")
df_active = df[df["UNFULLFILMENT"] == "NO"].copy().reset_index(drop=True)
print(f"  A scorar: {len(df_active)}")

X_score = np.array([encode_row(r) for _, r in df_active.iterrows()])
probs   = pipe.predict_proba(X_score)[:, 1]

df_active["risk_score"] = probs
df_active["risk_pct"]   = (probs * 100).round(1)
# Umbrales calibrados: ALTO ~top 10%, MEDIO ~top 25%, BAJO resto
# Con modelo no-balanceado: >15% = ALTO, 8-15% = MEDIO, <8% = BAJO
df_active["risk_level"] = pd.cut(
    probs, bins=[-np.inf, 0.08, 0.15, np.inf],
    labels=["BAJO", "MEDIO", "ALTO"]
)
print(df_active["risk_level"].value_counts().to_string())

_MARCA_MAP = {"MASTERCARD": "Mastercard", "AMEX": "Amex", "CABAL": "Cabal"}

def risk_factors(r) -> list:
    facts = []
    tc    = strip_accents(str(r.get("Tipo de Tarjeta", "")))
    marca = str(r.get("Marca de Tarjeta", ""))
    if tc == "Debito": facts.append("Tarjeta debito")
    if marca in _MARCA_MAP: facts.append(_MARCA_MAP[marca])
    edad = str(r.get("Rango etario", ""))
    if edad == "25 a 34": facts.append("Edad 25-34")
    if edad == "< 25":    facts.append("< 25 anios")
    if str(r.get("Email",   "")).upper() != "SI": facts.append("Sin email")
    if str(r.get("Celular", "")).upper() != "SI": facts.append("Sin celular")
    return facts

df_active["risk_factors"] = [risk_factors(r) for _, r in df_active.iterrows()]

# ── 6. ESTADISTICAS POR FACER ─────────────────────────────────────────────────
print("Calculando estadisticas por facer...")
facer_grp = df_active.groupby("Facer")
facer_df  = facer_grp.agg(
    donantes      =("risk_score", "count"),
    score_promedio=("risk_score", "mean"),
    alto_riesgo   =("risk_level", lambda x: (x == "ALTO").sum()),
    medio_riesgo  =("risk_level", lambda x: (x == "MEDIO").sum()),
    proj_bajas    =("risk_score", "sum"),
).reset_index()

facer_df["pct_alto"]       = (facer_df["alto_riesgo"] / facer_df["donantes"] * 100).round(1)
facer_df["score_prom_pct"] = (facer_df["score_promedio"] * 100).round(1)
facer_df["proj_bajas"]     = facer_df["proj_bajas"].round(1)

# Historial UF por facer (via username lookup)
hist_rates, hist_totals, hist_ufs = [], [], []
for _, row in facer_df.iterrows():
    mask = df_active["Facer"] == row["Facer"]
    usuarios = df_active.loc[mask, "Usuario"].values
    if len(usuarios) > 0:
        u = str(usuarios[0]).strip()
        s = facer_stats[u]
        hist_rates.append(round(s["uf"] / s["t"] * 100, 1) if s["t"] > 0 else 0.0)
        hist_totals.append(s["t"])
        hist_ufs.append(s["uf"])
    else:
        hist_rates.append(0.0); hist_totals.append(0); hist_ufs.append(0)

facer_df["hist_uf_rate"] = hist_rates
facer_df["hist_total"]   = hist_totals
facer_df = facer_df.sort_values("score_promedio", ascending=False)

# ── 7. ESTADISTICAS POR REGION ────────────────────────────────────────────────
region_df = df_active.groupby("IMPUTACION REGION").agg(  # column already normalized
    donantes      =("risk_score", "count"),
    score_promedio=("risk_score", "mean"),
    alto          =("risk_level", lambda x: (x == "ALTO").sum()),
).reset_index()
region_df["score_pct"] = (region_df["score_promedio"] * 100).round(1)
region_df = region_df.sort_values("score_promedio", ascending=False)

# ── 8. SERIALIZAR ─────────────────────────────────────────────────────────────
print("Serializando datos para HTML...")

# Contar facers activos en la cartera actual
facers_activos_cnt = sum(
    1 for fn, u in {r.get("Facer", ""): r.get("Usuario", "") for _, r in df_active.iterrows()}.items()
    if facer_status.get(str(u).strip().lower(), {}).get("estado") == "ACTIVO"
)

summary = {
    "total_active":       len(df_active),
    "alto":               int((df_active["risk_level"] == "ALTO").sum()),
    "medio":              int((df_active["risk_level"] == "MEDIO").sum()),
    "bajo":               int((df_active["risk_level"] == "BAJO").sum()),
    "proj_bajas_3m":      round(float(df_active["risk_score"].sum()), 1),
    "cv_auc":             round(float(cv_auc.mean()), 3),
    "train_auc":          round(float(train_auc), 3),
    "train_size":         int(len(df_train)),
    "train_pos":          int(pos),
    "generated_at":       datetime.now().strftime("%d-%b-%Y %H:%M"),
    "global_uf_rate":     round(GLOBAL_UF_RATE * 100, 2),
    "meses_entrenamiento":"DIC-2025 + ENE-2026",
    "facers_activos":     facers_activos_cnt,
    "facers_total":       len(facer_df),
}

def _facer_nomina(usuario: str) -> dict:
    info = facer_status.get(str(usuario).strip().lower(), {})
    return {
        "estado_nomina": info.get("estado", "SIN INFO"),
        "puesto_nomina": info.get("puesto", ""),
        "meses_prov":    info.get("meses_prov", []),
    }

donors_data = []
for _, r in df_active.sort_values("risk_score", ascending=False).iterrows():
    mn  = float(r.get("monto_num", 0) or 0)
    nom = _facer_nomina(r.get("Usuario", ""))
    donors_data.append({
        "nombre":        f"{r.get('Nombre donante','')} {r.get('Apellido donante','')}".strip(),
        "facer":         r.get("Facer",   ""),
        "usuario":       r.get("Usuario", ""),
        "region":        r.get("IMPUTACION REGION",  ""),
        "ciudad":        r.get("IMPUTACION CIUDAD",  ""),
        "mes_alta":      r.get("MES-ANO", ""),
        "monto":         int(mn) if mn > 0 else None,
        "tarjeta":       f"{r.get('Tipo de Tarjeta','')} {r.get('Marca de Tarjeta','')}".strip(),
        "edad":          r.get("Rango etario", ""),
        "score":         round(float(r["risk_score"]) * 100, 1),
        "level":         str(r["risk_level"]),
        "factors":       list(r["risk_factors"]),
        "estado_nomina": nom["estado_nomina"],
    })

facers_data = []
for _, r in facer_df.iterrows():
    facer_name = str(r["Facer"]).strip()
    # look up username for this facer name from df_active
    mask = df_active["Facer"] == facer_name
    usuario = df_active.loc[mask, "Usuario"].iloc[0] if mask.any() else ""
    nom = _facer_nomina(usuario)
    facers_data.append({
        "facer":         facer_name,
        "usuario":       usuario,
        "donantes":      int(r["donantes"]),
        "score_pct":     float(r["score_prom_pct"]),
        "pct_alto":      float(r["pct_alto"]),
        "alto":          int(r["alto_riesgo"]),
        "medio":         int(r["medio_riesgo"]),
        "proj_bajas":    float(r["proj_bajas"]),
        "hist_uf_rate":  float(r["hist_uf_rate"]),
        "hist_total":    int(r["hist_total"]),
        "estado_nomina": nom["estado_nomina"],
        "puesto_nomina": nom["puesto_nomina"],
        "meses_prov":    nom["meses_prov"],
    })

regions_data = []
for _, r in region_df.iterrows():
    regions_data.append({
        "region":    str(r["IMPUTACION REGION"]).strip() or "(Sin region)",
        "donantes":  int(r["donantes"]),
        "score_pct": float(r["score_pct"]),
        "alto":      int(r["alto"]),
    })

feat_importance = [
    {
        "feature": NOMBRE_FEAT.get(n, n),
        "coef":    round(float(c), 3),
        "abs":     round(abs(float(c)), 3),
    }
    for n, c in importances[:12]
]

payload = json.dumps({
    "summary":  summary,
    "donors":   donors_data,
    "facers":   facers_data,
    "regions":  regions_data,
    "features": feat_importance,
}, ensure_ascii=False)

# ── 9. HTML ───────────────────────────────────────────────────────────────────
print("Generando HTML...")

HTML_TEMPLATE = """<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Predictivo UF · UNICEF Argentina 2026</title>
<style>
:root {
  --bg:#0f1117; --surface:#1a1d2e; --surface2:#242740; --border:#2e3250;
  --text:#e2e4f0; --muted:#7b80a0; --alto:#ff4d6d; --alto-bg:rgba(255,77,109,.12);
  --medio:#fbbf24; --medio-bg:rgba(251,191,36,.12); --bajo:#34d399; --bajo-bg:rgba(52,211,153,.12);
  --accent:#6366f1; --accent2:#818cf8;
}
*{box-sizing:border-box;margin:0;padding:0;}
body{background:var(--bg);color:var(--text);font-family:'Segoe UI',system-ui,sans-serif;font-size:14px;}
header{background:linear-gradient(135deg,#1a1d2e,#242740);border-bottom:1px solid var(--border);
  padding:16px 28px;display:flex;align-items:center;gap:14px;}
.logo{width:38px;height:38px;background:var(--accent);border-radius:10px;display:flex;
  align-items:center;justify-content:center;font-size:18px;flex-shrink:0;}
.title{font-size:17px;font-weight:700;color:#fff;}
.subtitle{font-size:11px;color:var(--muted);margin-top:2px;}
.meta{margin-left:auto;text-align:right;font-size:11px;color:var(--muted);line-height:1.5;}
main{padding:22px 28px;}
.cards{display:grid;grid-template-columns:repeat(5,1fr);gap:10px;margin-bottom:24px;}
.card{background:var(--surface);border:1px solid var(--border);border-radius:12px;padding:16px;}
.card-val{font-size:26px;font-weight:800;margin-bottom:3px;}
.card-label{font-size:10px;color:var(--muted);text-transform:uppercase;letter-spacing:.5px;}
.card-sub{font-size:11px;color:var(--muted);margin-top:5px;}
.card.alto .card-val{color:var(--alto);}
.card.medio .card-val{color:var(--medio);}
.card.bajo .card-val{color:var(--bajo);}
.card.proj .card-val{color:var(--accent2);}
.card-criteria{font-size:10px;color:var(--muted);margin-top:8px;border-top:1px solid var(--border);padding-top:7px;line-height:1.7;}
.card-criteria b{color:var(--text);font-weight:600;}
.card-crit-row{display:flex;gap:4px;align-items:flex-start;margin-top:3px;}
.card-crit-dot{width:6px;height:6px;border-radius:50%;flex-shrink:0;margin-top:3px;}
.card.alto .card-crit-dot{background:var(--alto);}
.card.medio .card-crit-dot{background:var(--medio);}
.card.bajo .card-crit-dot{background:var(--bajo);}
.nom-badge{display:inline-block;padding:2px 7px;border-radius:4px;font-size:10px;font-weight:700;}
.nom-ACTIVO{background:rgba(52,211,153,.15);color:#34d399;}
.nom-BAJA{background:rgba(255,77,109,.12);color:#ff4d6d;}
.nom-SIN{background:rgba(123,128,160,.12);color:#7b80a0;}
.toggle-btn{background:var(--surface2);border:1px solid var(--border);color:var(--muted);
  padding:5px 12px;border-radius:20px;cursor:pointer;font-size:11px;font-weight:600;transition:all .15s;}
.toggle-btn.on{background:rgba(52,211,153,.15);border-color:#34d399;color:#34d399;}
.tabs{display:flex;gap:8px;margin-bottom:20px;}
.tab-btn{background:var(--surface);border:1px solid var(--border);color:var(--muted);
  padding:7px 16px;border-radius:8px;cursor:pointer;font-size:13px;font-weight:600;transition:all .15s;}
.tab-btn:hover{color:var(--text);border-color:var(--accent);}
.tab-btn.active{background:var(--accent);border-color:var(--accent);color:#fff;}
.section-title{font-size:15px;font-weight:700;color:#fff;margin-bottom:12px;}
.section-sub{font-size:11px;color:var(--muted);margin-top:-8px;margin-bottom:12px;}
.table-wrap{overflow-x:auto;border-radius:10px;border:1px solid var(--border);}
table{width:100%;border-collapse:collapse;}
thead th{background:var(--surface2);padding:9px 13px;text-align:left;font-size:10px;font-weight:600;
  text-transform:uppercase;letter-spacing:.5px;color:var(--muted);border-bottom:1px solid var(--border);
  white-space:nowrap;cursor:pointer;user-select:none;}
thead th:hover{color:var(--text);}
thead th.sorted{color:var(--accent2);}
tbody tr{border-bottom:1px solid var(--border);transition:background .12s;}
tbody tr:last-child{border-bottom:none;}
tbody tr:hover{background:var(--surface2);}
tbody td{padding:9px 13px;vertical-align:middle;}
.risk{display:inline-flex;align-items:center;gap:4px;padding:2px 9px;border-radius:20px;
  font-size:10px;font-weight:700;}
.risk.ALTO{background:var(--alto-bg);color:var(--alto);}
.risk.MEDIO{background:var(--medio-bg);color:var(--medio);}
.risk.BAJO{background:var(--bajo-bg);color:var(--bajo);}
.dot{width:5px;height:5px;border-radius:50%;flex-shrink:0;}
.ALTO .dot{background:var(--alto);}
.MEDIO .dot{background:var(--medio);}
.BAJO .dot{background:var(--bajo);}
.score-wrap{display:flex;align-items:center;gap:7px;min-width:110px;}
.bar-bg{flex:1;height:5px;background:var(--surface2);border-radius:3px;overflow:hidden;}
.bar{height:100%;border-radius:3px;}
.snum{font-size:12px;font-weight:700;min-width:36px;text-align:right;}
.factors{display:flex;flex-wrap:wrap;gap:3px;}
.factor{padding:2px 6px;background:var(--surface2);border-radius:4px;font-size:10px;color:var(--muted);}
.filters{display:flex;gap:8px;margin-bottom:10px;flex-wrap:wrap;align-items:center;}
.filters input,.filters select{background:var(--surface);border:1px solid var(--border);color:var(--text);
  border-radius:8px;padding:6px 11px;font-size:13px;outline:none;transition:border-color .15s;}
.filters input:focus,.filters select:focus{border-color:var(--accent);}
.filters input{width:210px;}
.pager{display:flex;align-items:center;gap:8px;margin-top:10px;justify-content:flex-end;
  font-size:11px;color:var(--muted);}
.pager button{background:var(--surface2);border:1px solid var(--border);color:var(--text);
  border-radius:6px;padding:4px 11px;cursor:pointer;font-size:12px;}
.pager button:disabled{opacity:.3;cursor:default;}
.feat-grid{display:grid;grid-template-columns:1fr 1fr;gap:7px;}
.feat-item{display:flex;align-items:center;gap:9px;padding:8px 11px;
  background:var(--surface);border-radius:8px;}
.feat-label{flex:1;font-size:12px;}
.feat-bar-wrap{width:75px;height:5px;background:var(--surface2);border-radius:3px;overflow:hidden;}
.feat-bar{height:100%;border-radius:3px;}
.feat-coef{font-size:11px;font-weight:700;min-width:44px;text-align:right;}
.pos{color:var(--alto);} .neg{color:var(--bajo);}
.metric-grid{display:grid;grid-template-columns:repeat(3,1fr);gap:16px;text-align:center;
  background:var(--surface);border:1px solid var(--border);border-radius:10px;padding:18px;margin-top:18px;}
.metric-val{font-size:22px;font-weight:800;color:var(--accent2);}
.metric-label{font-size:10px;color:var(--muted);margin-top:3px;}
.model-info{background:var(--surface);border:1px solid var(--border);border-radius:10px;
  padding:18px;line-height:1.7;color:var(--muted);margin-bottom:18px;}
.model-info strong{color:var(--text);}
.exp-btn{background:var(--accent);border:none;color:#fff;padding:6px 13px;border-radius:7px;
  cursor:pointer;font-size:12px;font-weight:600;}
.exp-btn:hover{opacity:.9;}
@media(max-width:900px){
  main{padding:14px;} .cards{grid-template-columns:repeat(2,1fr);}
  .feat-grid{grid-template-columns:1fr;}
}
</style>
</head>
<body>
<header>
  <div class="logo">&#127919;</div>
  <div>
    <div class="title">Predictivo Unfullfilment &middot; UNICEF Argentina 2026</div>
    <div class="subtitle">Scoring de donantes activos &middot; Horizonte 3 meses</div>
  </div>
  <div class="meta" id="meta-info"></div>
</header>
<main>
  <div class="cards" id="cards"></div>
  <div class="tabs">
    <button class="tab-btn active" onclick="showTab('tab-don')">Donantes</button>
    <button class="tab-btn" onclick="showTab('tab-fac')">Facers</button>
    <button class="tab-btn" onclick="showTab('tab-reg')">Regiones</button>
    <button class="tab-btn" onclick="showTab('tab-mod')">Modelo</button>
  </div>

  <!-- DONANTES -->
  <div id="tab-don" class="tab-pane">
    <div class="section-title">Donantes activos &middot; Riesgo de baja (3 meses)</div>
    <div class="section-sub">Donantes con UNFULLFILMENT=NO &middot; click en columna para ordenar</div>
    <div class="filters">
      <input id="srch" type="text" placeholder="Buscar donante o facer..." oninput="filt()">
      <select id="flv" onchange="filt()">
        <option value="">Todos los niveles</option>
        <option value="ALTO">Rojo - Alto riesgo</option>
        <option value="MEDIO">Amarillo - Medio riesgo</option>
        <option value="BAJO">Verde - Bajo riesgo</option>
      </select>
      <select id="ffc" onchange="filt()"><option value="">Todos los facers</option></select>
      <select id="frg" onchange="filt()"><option value="">Todas las regiones</option></select>
      <select id="fms" onchange="filt()"><option value="">Todos los meses</option></select>
      <button id="btn-activos" class="toggle-btn on" onclick="toggleActivos()" title="Mostrar solo facers activos en nomina">&#9679; Solo activos en nomina</button>
      <button class="exp-btn" onclick="exportCSV()">&#8595; CSV</button>
      <span id="dcount" style="margin-left:auto;font-size:11px;color:var(--muted);"></span>
    </div>
    <div class="table-wrap">
      <table>
        <thead><tr>
          <th onclick="srt('d','nombre')">Donante</th>
          <th onclick="srt('d','facer')">Facer</th>
          <th onclick="srt('d','estado_nomina')">Nomina</th>
          <th onclick="srt('d','region')">Region</th>
          <th onclick="srt('d','mes_alta')">Mes alta</th>
          <th onclick="srt('d','monto')">Monto</th>
          <th onclick="srt('d','tarjeta')">Tarjeta</th>
          <th onclick="srt('d','edad')">Edad</th>
          <th onclick="srt('d','score')" class="sorted" id="th-score">Score &#8595;</th>
          <th>Factores de riesgo</th>
        </tr></thead>
        <tbody id="dbody"></tbody>
      </table>
    </div>
    <div class="pager">
      <span id="pginfo"></span>
      <button id="bprev" onclick="pg(-1)">&#8592; Ant</button>
      <button id="bnext" onclick="pg(1)">Sig &#8594;</button>
    </div>
  </div>

  <!-- FACERS -->
  <div id="tab-fac" class="tab-pane" style="display:none">
    <div class="section-title">Ranking de Facers por riesgo de baja</div>
    <div class="section-sub">Ordenado por score promedio de su cartera activa</div>
    <div class="filters">
      <input id="srchf" type="text" placeholder="Buscar facer..." oninput="rndF()">
      <button id="btn-fac-activos" class="toggle-btn on" onclick="toggleFacActivos()" title="Mostrar solo activos en nomina">&#9679; Solo activos en nomina</button>
    </div>
    <div class="table-wrap">
      <table>
        <thead><tr>
          <th>#</th>
          <th onclick="srt('f','facer')">Facer</th>
          <th onclick="srt('f','donantes')">Donantes</th>
          <th onclick="srt('f','pct_alto')">% Alto riesgo</th>
          <th onclick="srt('f','score_pct')" class="sorted">Score prom &#8595;</th>
          <th onclick="srt('f','proj_bajas')">Proy. bajas 3m</th>
          <th onclick="srt('f','hist_uf_rate')">UF historico</th>
          <th onclick="srt('f','estado_nomina')">Nomina</th>
        </tr></thead>
        <tbody id="fbody"></tbody>
      </table>
    </div>
  </div>

  <!-- REGIONES -->
  <div id="tab-reg" class="tab-pane" style="display:none">
    <div class="section-title">Riesgo por region</div>
    <div class="table-wrap">
      <table>
        <thead><tr>
          <th>Region</th><th>Donantes</th><th>Alto riesgo</th>
          <th>Score prom</th><th style="width:180px">Distribucion</th>
        </tr></thead>
        <tbody id="rbody"></tbody>
      </table>
    </div>
  </div>

  <!-- MODELO -->
  <div id="tab-mod" class="tab-pane" style="display:none">
    <div class="section-title">Metodologia del modelo predictivo</div>
    <div class="model-info" id="modinfo"></div>
    <div class="section-title">Features mas influyentes</div>
    <div class="feat-grid" id="featgrid"></div>
    <div class="metric-grid" id="mmetrics"></div>
  </div>
</main>

<script>
const DATA = __PAYLOAD__;

// ─ helpers ─
const $ = id => document.getElementById(id);
const fmt = n => n ? '$' + n.toLocaleString('es-AR') : '-';
const sc  = s => s >= 15 ? 'var(--alto)' : s >= 7 ? 'var(--medio)' : 'var(--bajo)';

let dSorted = [...DATA.donors];
let fSorted = [...DATA.facers];
let dPage   = 0;
const PG    = 50;
let soloActivos    = true;
let soloActivosFac = true;

const nomBadge = e => {
  if (e === 'ACTIVO')   return '<span class="nom-badge nom-ACTIVO">ACTIVO</span>';
  if (e === 'BAJA')     return '<span class="nom-badge nom-BAJA">BAJA</span>';
  return '<span class="nom-badge nom-SIN">S/D</span>';
};

// ─ init ─
(function init() {
  const s = DATA.summary;
  $('meta-info').innerHTML =
    `Generado: ${s.generated_at}<br>Entrenado con ${s.train_size} donantes (${s.meses_entrenamiento})`;

  // factores positivos (aumentan riesgo) y protectores (reducen riesgo) del modelo
  const riskFeat  = DATA.features.filter(f=>f.coef>0).slice(0,3).map(f=>f.feature);
  const safeFeat  = DATA.features.filter(f=>f.coef<0).slice(0,3).map(f=>f.feature);
  const mkFactors = (arr, cls) => arr.map(f=>`<div class="card-crit-row"><div class="card-crit-dot ${cls}"></div><span>${f}</span></div>`).join('');

  $('cards').innerHTML = `
    <div class="card">
      <div class="card-val">${s.total_active.toLocaleString('es-AR')}</div>
      <div class="card-label">Donantes activos</div>
      <div class="card-sub">UNFULLFILMENT = NO</div>
      <div class="card-criteria"><b>Tasa UF base:</b> ${s.global_uf_rate}%<br>Entrenado: ${s.meses_entrenamiento}</div>
    </div>
    <div class="card alto">
      <div class="card-val">${s.alto.toLocaleString('es-AR')}</div>
      <div class="card-label">&#128308; Alto riesgo</div>
      <div class="card-sub">${((s.alto/s.total_active)*100).toFixed(1)}% del total</div>
      <div class="card-criteria">
        <b>Criterio:</b> probabilidad de baja &ge; 15%
        <div style="margin-top:5px;color:var(--alto);font-weight:600;font-size:9px;text-transform:uppercase;letter-spacing:.4px;">Factores que elevan el riesgo</div>
        ${mkFactors(riskFeat,'alto')}
      </div>
    </div>
    <div class="card medio">
      <div class="card-val">${s.medio.toLocaleString('es-AR')}</div>
      <div class="card-label">&#128993; Medio riesgo</div>
      <div class="card-sub">${((s.medio/s.total_active)*100).toFixed(1)}% del total</div>
      <div class="card-criteria">
        <b>Criterio:</b> probabilidad de baja entre 8% y 15%
        <div style="margin-top:5px;color:var(--medio);font-weight:600;font-size:9px;text-transform:uppercase;letter-spacing:.4px;">Riesgo moderado</div>
        ${mkFactors(riskFeat.slice(0,2),'medio')}
      </div>
    </div>
    <div class="card bajo">
      <div class="card-val">${s.bajo.toLocaleString('es-AR')}</div>
      <div class="card-label">&#128994; Bajo riesgo</div>
      <div class="card-sub">${((s.bajo/s.total_active)*100).toFixed(1)}% del total</div>
      <div class="card-criteria">
        <b>Criterio:</b> probabilidad de baja &lt; 8%
        <div style="margin-top:5px;color:var(--bajo);font-weight:600;font-size:9px;text-transform:uppercase;letter-spacing:.4px;">Factores protectores</div>
        ${mkFactors(safeFeat,'bajo')}
      </div>
    </div>
    <div class="card proj">
      <div class="card-val">~${Math.round(s.proj_bajas_3m)}</div>
      <div class="card-label">&#128200; Proy. bajas 3m</div>
      <div class="card-sub">${s.facers_activos} / ${s.facers_total} facers activos</div>
      <div class="card-criteria"><b>Metodo:</b> suma de probabilidades individuales de cada donante activo</div>
    </div>`;

  // fill selects (solo facers activos por defecto)
  const uFacers = [...new Set(DATA.donors.filter(d=>d.estado_nomina==='ACTIVO').map(d => d.facer).filter(Boolean))].sort();
  uFacers.forEach(f => { const o = new Option(f, f); $('ffc').add(o); });
  const uReg = [...new Set(DATA.donors.map(d => d.region).filter(Boolean))].sort();
  uReg.forEach(r => { const o = new Option(r, r); $('frg').add(o); });
  const uMes = [...new Set(DATA.donors.map(d => d.mes_alta).filter(Boolean))].sort();
  uMes.forEach(m => { const o = new Option(m, m); $('fms').add(o); });

  filt(); rndF(); rndR(); rndM();
})();

// ─ tabs ─
function showTab(id) {
  document.querySelectorAll('.tab-pane').forEach(el => el.style.display = 'none');
  $(id).style.display = 'block';
  document.querySelectorAll('.tab-btn').forEach(btn =>
    btn.classList.toggle('active', btn.getAttribute('onclick').includes(id)));
}

// ─ toggles nomina ─
function toggleActivos() {
  soloActivos = !soloActivos;
  $('btn-activos').classList.toggle('on', soloActivos);
  $('btn-activos').textContent = soloActivos ? '● Solo activos en nomina' : '○ Todos los facers';
  dPage = 0; filt();
}
function toggleFacActivos() {
  soloActivosFac = !soloActivosFac;
  $('btn-fac-activos').classList.toggle('on', soloActivosFac);
  $('btn-fac-activos').textContent = soloActivosFac ? '● Solo activos en nomina' : '○ Todos los facers';
  rndF();
}

// ─ donors ─
function filt() {
  const q  = $('srch').value.toLowerCase();
  const lv = $('flv').value;
  const fc = $('ffc').value;
  const rg = $('frg').value;
  const ms = $('fms').value;
  dSorted = DATA.donors.filter(d =>
    (!soloActivos || d.estado_nomina === 'ACTIVO') &&
    (!q  || d.nombre.toLowerCase().includes(q) || d.facer.toLowerCase().includes(q)) &&
    (!lv || d.level  === lv) &&
    (!fc || d.facer  === fc) &&
    (!rg || d.region === rg) &&
    (!ms || d.mes_alta === ms));
  dPage = 0; rndD();
}

function rndD() {
  const start = dPage * PG;
  const page  = dSorted.slice(start, start + PG);
  $('dcount').textContent = dSorted.length.toLocaleString('es-AR') + ' donantes';
  $('dbody').innerHTML = page.map(d => `
    <tr>
      <td style="font-weight:600">${d.nombre||'-'}</td>
      <td style="font-size:12px;color:var(--muted)">${d.facer||'-'}</td>
      <td>${nomBadge(d.estado_nomina)}</td>
      <td style="font-size:12px">${d.region||'-'}</td>
      <td style="font-size:12px">${d.mes_alta||'-'}</td>
      <td style="font-size:12px">${fmt(d.monto)}</td>
      <td style="font-size:12px">${d.tarjeta||'-'}</td>
      <td style="font-size:12px">${d.edad||'-'}</td>
      <td>
        <div class="score-wrap">
          <div class="bar-bg"><div class="bar" style="width:${Math.min(d.score,100)}%;background:${sc(d.score)}"></div></div>
          <div class="snum" style="color:${sc(d.score)}">${d.score}%</div>
        </div>
        <div style="margin-top:3px"><span class="risk ${d.level}"><span class="dot"></span>${d.level}</span></div>
      </td>
      <td><div class="factors">${(d.factors||[]).map(f=>`<span class="factor">${f}</span>`).join('')}</div></td>
    </tr>`).join('');
  const tp = Math.ceil(dSorted.length / PG);
  $('pginfo').textContent = `Pag ${dPage+1}/${tp} · ${dSorted.length.toLocaleString('es-AR')} filas`;
  $('bprev').disabled = dPage === 0;
  $('bnext').disabled = dPage >= tp - 1;
}

function pg(d) { dPage += d; rndD(); }

// ─ sort ─
const SS = { d: {col:'score',dir:-1}, f: {col:'score_pct',dir:-1} };
function srt(tbl, col) {
  if (SS[tbl].col === col) SS[tbl].dir *= -1;
  else { SS[tbl].col = col; SS[tbl].dir = -1; }
  const arr = tbl === 'd' ? dSorted : fSorted;
  arr.sort((a,b) => {
    const av = a[col], bv = b[col];
    if (typeof av === 'number') return (av-bv) * SS[tbl].dir;
    return String(av).localeCompare(String(bv),'es') * SS[tbl].dir;
  });
  if (tbl === 'd') { dPage = 0; rndD(); }
  else rndF();
}

// ─ facers ─
function rndF() {
  const q = ($('srchf')?.value||'').toLowerCase();
  const data = fSorted.filter(f =>
    (!soloActivosFac || f.estado_nomina === 'ACTIVO') &&
    (!q || f.facer.toLowerCase().includes(q))
  );
  const mx = Math.max(...data.map(f=>f.score_pct), 1);
  $('fbody').innerHTML = data.map((f,i) => `
    <tr>
      <td style="color:var(--muted);font-weight:700">${i+1}</td>
      <td style="font-weight:600">${f.facer}${f.puesto_nomina?'<br><span style="font-size:10px;color:var(--muted);font-weight:400">'+f.puesto_nomina+'</span>':''}</td>
      <td>${f.donantes}</td>
      <td><span style="color:${f.pct_alto>20?'var(--alto)':f.pct_alto>10?'var(--medio)':'var(--bajo)'};font-weight:700">${f.pct_alto}%</span>
          <span style="color:var(--muted);font-size:11px"> (${f.alto})</span></td>
      <td>
        <div style="display:flex;align-items:center;gap:7px">
          <div style="width:90px;background:var(--surface2);border-radius:4px;overflow:hidden;height:7px">
            <div style="width:${(f.score_pct/mx*100).toFixed(0)}%;height:100%;background:var(--accent);border-radius:4px"></div>
          </div>
          <strong>${f.score_pct}%</strong>
        </div>
      </td>
      <td style="font-weight:700;color:var(--alto)">~${Math.ceil(f.proj_bajas)}</td>
      <td style="font-size:12px;color:var(--muted)">${f.hist_uf_rate>0?f.hist_uf_rate+'% ('+f.hist_total+' don.)':'-'}</td>
      <td>${nomBadge(f.estado_nomina)}${f.meses_prov&&f.meses_prov.length?'<br><span style="font-size:9px;color:var(--muted)">'+f.meses_prov.join(', ')+'</span>':''}</td>
    </tr>`).join('');
}

// ─ regions ─
function rndR() {
  const mx = Math.max(...DATA.regions.map(r=>r.score_pct),1);
  $('rbody').innerHTML = DATA.regions.map(r => `
    <tr>
      <td style="font-weight:600">${r.region}</td>
      <td>${r.donantes}</td>
      <td><span style="color:var(--alto);font-weight:700">${r.alto}</span>
          <span style="color:var(--muted);font-size:11px"> (${r.donantes>0?((r.alto/r.donantes)*100).toFixed(1):0}%)</span></td>
      <td><strong>${r.score_pct}%</strong></td>
      <td><div style="background:var(--surface2);border-radius:4px;overflow:hidden;height:9px">
            <div style="width:${(r.score_pct/mx*100).toFixed(0)}%;height:100%;background:var(--accent);border-radius:4px"></div>
          </div></td>
    </tr>`).join('');
}

// ─ model ─
function rndM() {
  const s = DATA.summary;
  $('modinfo').innerHTML = `
    <p><strong>Tipo de modelo:</strong> Regresion Logistica con regularizacion L2 (C=0.5), pesos balanceados por clase.</p>
    <p style="margin-top:7px"><strong>Entrenamiento:</strong> Donantes de DIC-2025 y ENE-2026 (${s.train_size.toLocaleString()} registros, ${s.train_pos} UF=SI). Esos meses ya tienen tiempo suficiente para que los estados UF esten definidos.</p>
    <p style="margin-top:7px"><strong>Variable target:</strong> UNFULLFILMENT = SI (cualquier tipo: rechazo reiterado, problemas economicos, no contactado, etc.).</p>
    <p style="margin-top:7px"><strong>Aplicacion:</strong> Scored sobre los donantes activos con UNFULLFILMENT=NO. El score representa la probabilidad estimada de entrar en proceso UF en los proximos ~3 meses.</p>
    <p style="margin-top:7px"><strong>Niveles:</strong> ALTO &ge;15% | MEDIO 8-15% | BAJO &lt;8%</p>`;

  const mx = Math.max(...DATA.features.map(f=>f.abs), .01);
  $('featgrid').innerHTML = DATA.features.map(f => `
    <div class="feat-item">
      <div class="feat-label">${f.feature}</div>
      <div class="feat-bar-wrap"><div class="feat-bar" style="width:${(f.abs/mx*100).toFixed(0)}%;background:${f.coef>0?'var(--alto)':'var(--bajo)'}"></div></div>
      <div class="feat-coef ${f.coef>0?'pos':'neg'}">${f.coef>0?'+':''}${f.coef}</div>
    </div>`).join('');

  $('mmetrics').innerHTML = `
    <div><div class="metric-val">${s.cv_auc}</div><div class="metric-label">AUC Validacion Cruzada (5-fold)</div></div>
    <div><div class="metric-val">${s.train_auc}</div><div class="metric-label">AUC en entrenamiento</div></div>
    <div><div class="metric-val">${s.global_uf_rate}%</div><div class="metric-label">Tasa UF base en entrenamiento</div></div>`;
}

// ─ export ─
function exportCSV() {
  const hdr = ['Donante','Facer','Nomina','Region','Ciudad','Mes Alta','Monto','Tarjeta','Edad','Score (%)','Nivel','Factores'];
  const rows = dSorted.map(d => [
    d.nombre, d.facer, d.estado_nomina, d.region, d.ciudad, d.mes_alta,
    d.monto||'', d.tarjeta, d.edad, d.score, d.level,
    (d.factors||[]).join(' | ')
  ]);
  const csv = [hdr,...rows].map(r=>r.map(v=>`"${String(v).replace(/"/g,'""')}"`).join(',')).join('\\n');
  const a = document.createElement('a');
  a.href = URL.createObjectURL(new Blob(['\\uFEFF'+csv],{type:'text/csv;charset=utf-8'}));
  a.download = 'predictivo_uf_' + new Date().toISOString().slice(0,10) + '.csv';
  a.click();
}
</script>
</body>
</html>"""

html_out = HTML_TEMPLATE.replace("__PAYLOAD__", payload)
with open(OUTPUT, "w", encoding="utf-8") as f:
    f.write(html_out)

print(f"\nDashboard generado: {OUTPUT}")
print(f"  Activos scoreados:  {len(df_active)}")
print(f"  Alto riesgo:        {(df_active['risk_level']=='ALTO').sum()}")
print(f"  Proyeccion bajas:   ~{round(float(df_active['risk_score'].sum()))}")
